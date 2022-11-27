import os
from openpyxl.styles import Side, Border
from openpyxl import Workbook

from datetime import datetime
import pytz



def WRITE_HEADERS_TO_EXCEL(equity_dict, search_dict, vote_type, equity_total, all_temp_dicts, testing=False):
    
    my_time = pytz.timezone('US/Eastern') 
    current_datetime = datetime.now(my_time).replace(microsecond=0).replace(tzinfo=None)
    current_date = current_datetime.strftime('%Y-%m-%d')
    #print(current_datetime)
    # print(current_date)


    wb = Workbook()
    ws = wb.active

    #print("WRITE_HEADERS_TO_EXCEL========")
    #print("TESTING        :", testing)
    #print("VOTE TYPE      :", vote_type)
    #print("EQUITY AMOUNTS :", equity_total)
    #print("EQUITY DICT    :", equity_dict)
    #print()

    #TODO SHOW EXPENSIES

    # TITLE
    ws.cell(row=1, column=1).value = "MANSURA"
    ws.cell(row=1, column=2).value = str(current_date)
    ws.cell(row=1, column=3).value = vote_type

    # EXPENSES
    ws.cell(row=3, column=1).value = "EXPENSES"

    ws.cell(row=4, column=1).value = "SERVER"
    ws.cell(row=4, column=2).value = "DATABASE"
    ws.cell(row=4, column=3).value = "DOMAIN"
    ws.cell(row=4, column=4).value = "INC TRANS FEES"

    ws.cell(row=5, column=1).value = "$-x"
    ws.cell(row=5, column=2).value = "$-x"
    ws.cell(row=5, column=3).value = "$-x"
    ws.cell(row=5, column=4).value = "$-x"


    # TOTAL AMOUNT FOR THIS PERIOD
    ws.cell(row=7, column=1).value = "SYSEM TOT"
    ws.cell(row=7, column=2).value = "EQUITY TOT"
    ws.cell(row=7, column=3).value = "POOL TOT"
    ws.cell(row=8, column=1).value = f"${equity_total * 5}"
    ws.cell(row=8, column=2).value = f"${equity_total}"
    ws.cell(row=8, column=3).value = f"${(equity_total * 5 ) - equity_total}"


    # EQUITY DISTRIBUTION
    ws.cell(row=10, column=1).value = "EQUITY"  
    i = 1
    for key, value in equity_dict.items():
        # print(key, value)
        single_string = f'{key} ${"{:.2f}".format(float(value / 100) * equity_total)}'
        
        # print(single_string)
        # print(i, ":", single_string)
        ws.cell(row=11, column=i).value = single_string

        i+=1

    #USERS DISTRIBUTION
    ws.cell(row=13, column=1).value = "USERS"
    ws.cell(row=13, column=1).value = "FILE_ID"
    ws.cell(row=13, column=2).value = "VOTES"
    ws.cell(row=13, column=3).value = "PERCENT"
    ws.cell(row=13, column=4).value = "AMOUNT"
    ws.cell(row=13, column=5).value = "REPLYING_TO"
    ws.cell(row=13, column=6).value = "RATIO"
    ws.cell(row=13, column=7).value = "UPLOADER"
    ws.cell(row=13, column=8).value = "IN ORDER"

    ITERATION = 0
    for key, value in all_temp_dicts.items():
        # print(f"{key}:{value}")
        
        local_index = 0 # ANNOYING AND I NEED TO FIND A BETTER WAY TO DO THIS
        for key_, value_ in value["ORDER"].items():
            full_string = f"{key_} {value_}"

            if local_index == 0:
                ws.cell(row=14+ITERATION, column=7).value = full_string
            else:
                ws.cell(row=14+ITERATION, column=7+local_index).value = full_string
            local_index += 1
        

        # print(value)
        votes = f'{value["VALUE"]["VALUE"]["VOTES"]}'
        percent = f'{value["VALUE"]["VALUE"]["PERCENT"]}'
        amount = f'{value["VALUE"]["VALUE"]["AMOUNT"]}'

        if "RATIO" in value["VALUE"]:
            # print(value["VALUE"]["RATIO"])
            rat_key = (list(value["VALUE"]["RATIO"].keys())[0])
            rat_amount = value["VALUE"]["RATIO"][f"{rat_key}"]
            ratio_dict = str(rat_key) + " " + str(rat_amount)
        else:
            ratio_dict = ""
        
        if "REPLY" in value["VALUE"]:
            reply_to = value["VALUE"]["REPLY"]["REPLYING_TO"]
        else:
            reply_to = ""

        # ws.cell(row=14+ITERATION, column=8).value = "IN ORDER"
        ws.cell(row=14+ITERATION, column=1).value = key
        ws.cell(row=14+ITERATION, column=2).value = votes
        ws.cell(row=14+ITERATION, column=3).value = percent
        ws.cell(row=14+ITERATION, column=4).value = amount
        ws.cell(row=14+ITERATION, column=5).value = reply_to
        ws.cell(row=14+ITERATION, column=6).value = ratio_dict

        ITERATION +=1


    # print(F"TESTING IS: {testing}")
    wb.save(F"/root/mansura/Python/HISTORY/{vote_type}/{current_date}.xlsx")      

       
    